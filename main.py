from fastapi import FastAPI, UploadFile
from typing import Union, Tuple, List
from pydantic import BaseModel
from enum import Enum
import openpyxl
import difflib
from io import BytesIO
import pandas as pd
import re
import numpy as np

class Sides(Enum):
    top = 'Top'
    bottom = 'Bottom'


class Types(Enum):
    pressre = 'Pressure Pin'
    locating = 'Locating Pin'
    smd = 'SMD'
    hole = 'Through Hole'





class Test_Point(BaseModel):
    # pcb: PCB
    net: Union[int, None] = None
    name: str
    x_coord: float
    y_coord: float
    side: Sides
    type: Types
    hole_size: Union[float, None] = None


class PCB(BaseModel):
    # height: Tuple[float, Union[str, None]]
    # width: Tuple[float, Union[str, None]]
    # thickness: Tuple[float, Union[str, None]]
    height: float
    height_notes: Union[str, None]
    width: float
    width_notes: Union[str, None]
    thickness: float
    thickness_notes: Union[str, None]
    test_points: List[Test_Point] = []




app = FastAPI()


@app.get("/")
async def root():
    return "https://fixturfab.com/"


@app.post("/upload_pcb/")
async def upload_PCB(file: UploadFile):

    # print(BytesIO(file))
    xlsx = BytesIO(file.file.read())
    # workbook = openpyxl.load_workbook(xlsx)
    # PCB = workbook["PCB Specification"]
    # print(PCB)

    PCB_df = pd.read_excel(io = xlsx, sheet_name="PCB Specification", index_col=0, keep_default_na=False).transpose()
    test_point_df = pd.read_excel(io = xlsx, sheet_name="Test Point List", keep_default_na=False)
    # PCB.set_index("PCB Information",inplace = True)
    # print(PCB.loc("Height (mm)"))
    # print(PCB["PCB Information"].apply(lambda x: difflib.get_close_matches(x, "height")[0]))
    # for index, row in PCB.iterrows():
    #     if "height" in row[0].lower():
    #         print(row["Value"], row["Notes"])
    PCB_df.columns = [x.split(" ")[0].title() for x in PCB_df.columns.values]
    # t = PCB.set_index('PCB Information')[['Value', 'Notes']].T.apply(tuple).to_dict()
    Current_PCB = PCB(
        height=PCB_df["Height"]["Value"],
        height_notes=PCB_df["Height"]["Notes"],
        width=PCB_df["Width"]["Value"],
        width_notes=PCB_df["Width"]["Notes"],
        thickness=PCB_df["Thickness"]["Value"],
        thickness_notes=PCB_df["Thickness"]["Notes"]
    )

    test_point_df["Net"] = test_point_df["Net"].apply(lambda x: x.lstrip("NET"))
    test_point_df["X Coord"] = test_point_df["X Coord"].apply(lambda x: x.rstrip("mm"))
    test_point_df["Y Coord"] = test_point_df["Y Coord"].apply(lambda x: x.rstrip("mm"))
    test_point_df["Side"] = test_point_df["Side"].apply(lambda x: Sides(x))
    test_point_df["Type"] = test_point_df["Type"].apply(lambda x: Types(x))
    test_point_df["Hole Size"] = test_point_df["Hole Size"].apply(lambda x: x.rstrip("mm"))
    # print(test_point_df)
    test_point_df = test_point_df.apply(np.vectorize(lambda x: x if x else None))
    test_point_df.columns = [x.replace(" ", "_").lower() for x in test_point_df.columns.values]

    for iter, row in test_point_df.iterrows():
        point = row.to_dict()

        # point["pcb"] = Current_PCB
        p = Test_Point(**point)
        Current_PCB.test_points.append(p)

    return Current_PCB
# [0].apply(lambda x: difflib.get_close_matches(x, "height")[0]))


    # worksheet = wookbook.active
    #
    # # Iterate the loop to read the cell values
    # for i in range(0, worksheet.max_row):
    #     for col in worksheet.iter_cols(1, worksheet.max_column):
    #         print(col[i].value, end="\t\t")
    #     print('')
    return
